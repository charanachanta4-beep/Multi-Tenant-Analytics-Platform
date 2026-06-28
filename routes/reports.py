from flask import Blueprint, render_template
from flask_login import login_required, current_user
from sqlalchemy import func

from models.metric import Metric
from models.datasource import DataSource
from models import db

import csv

from openpyxl import Workbook
from openpyxl.styles import Font

from flask import send_file
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch

import io

reports_bp = Blueprint(
    "reports",
    __name__
)


@reports_bp.route("/reports")
@login_required
def reports():

    revenue = db.session.query(
        func.sum(Metric.revenue)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    visitors = db.session.query(
        func.sum(Metric.visitors)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    orders = db.session.query(
        func.sum(Metric.orders)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    profit = db.session.query(
        func.sum(Metric.profit)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    uploads = DataSource.query.filter_by(
    tenant_id=current_user.id
    ).count()

    average_revenue = (
        revenue / uploads
        if uploads else 0
    )

    average_profit = (
        profit / uploads
        if uploads else 0
    )

    best_day = (
        Metric.query.filter_by(
            tenant_id=current_user.id
        )
        .order_by(Metric.revenue.desc())
        .first()
    )

    return render_template(

        "reports.html",

        revenue=revenue,

        visitors=visitors,

        orders=orders,

        profit=profit,

        uploads=uploads,

        average_revenue=average_revenue,

        average_profit=average_profit,

        best_day=best_day

    )

@reports_bp.route("/reports/export/pdf")
@login_required
def export_pdf():

    revenue = db.session.query(
        func.sum(Metric.revenue)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    visitors = db.session.query(
        func.sum(Metric.visitors)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    orders = db.session.query(
        func.sum(Metric.orders)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    profit = db.session.query(
        func.sum(Metric.profit)
    ).filter(
        Metric.tenant_id == current_user.id
    ).scalar() or 0

    metrics = Metric.query.filter_by(
        tenant_id=current_user.id
    ).order_by(
        Metric.timestamp.desc()
    ).all()

    buffer = io.BytesIO()

    doc = SimpleDocTemplate(buffer)

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph("<b>BuildHub Analytics Report</b>", styles["Title"])
    )

    elements.append(
        Paragraph(
            f"Tenant : {current_user.company}",
            styles["Normal"]
        )
    )

    elements.append(
        Paragraph("<br/>", styles["Normal"])
    )

    summary = [

        ["Metric", "Value"],

        ["Revenue", f"₹{revenue:,.2f}"],

        ["Visitors", str(visitors)],

        ["Orders", str(orders)],

        ["Profit", f"₹{profit:,.2f}"]

    ]

    table = Table(summary, colWidths=[3*inch,3*inch])

    table.setStyle(TableStyle([

        ("BACKGROUND",(0,0),(-1,0),colors.black),

        ("TEXTCOLOR",(0,0),(-1,0),colors.white),

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,1),(-1,-1),colors.beige),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

        ("BOTTOMPADDING",(0,0),(-1,0),12),

    ]))

    elements.append(table)

    elements.append(
        Paragraph("<br/><b>Recent Metrics</b>", styles["Heading2"])
    )

    data = [

        ["Date","Revenue","Visitors","Orders","Profit"]

    ]

    for m in metrics[:15]:

        data.append([

            m.timestamp.strftime("%d-%m-%Y"),

            f"₹{m.revenue}",

            str(m.visitors),

            str(m.orders),

            f"₹{m.profit}"

        ])

    report = Table(data)

    report.setStyle(TableStyle([

        ("GRID",(0,0),(-1,-1),1,colors.black),

        ("BACKGROUND",(0,0),(-1,0),colors.lightgrey),

        ("ALIGN",(0,0),(-1,-1),"CENTER"),

    ]))

    elements.append(report)

    doc.build(elements)

    buffer.seek(0)

    return send_file(

        buffer,

        as_attachment=True,

        download_name="BuildHub_Report.pdf",

        mimetype="application/pdf"

    )

@reports_bp.route("/reports/export/excel")
@login_required
def export_excel():

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Analytics Report"

    headers = [
        "Date",
        "Revenue",
        "Visitors",
        "Orders",
        "Profit"
    ]

    for i, header in enumerate(headers, start=1):

        cell = sheet.cell(row=1, column=i)

        cell.value = header

        cell.font = Font(bold=True)

    metrics = Metric.query.filter_by(
        tenant_id=current_user.id
    ).order_by(
        Metric.timestamp.desc()
    ).all()

    row = 2

    for metric in metrics:

        sheet.cell(row=row, column=1).value = metric.timestamp.strftime("%d-%m-%Y")

        sheet.cell(row=row, column=2).value = metric.revenue

        sheet.cell(row=row, column=3).value = metric.visitors

        sheet.cell(row=row, column=4).value = metric.orders

        sheet.cell(row=row, column=5).value = metric.profit

        row += 1

    output = io.BytesIO()

    workbook.save(output)

    output.seek(0)

    return send_file(

        output,

        download_name="BuildHub_Report.xlsx",

        as_attachment=True,

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )

@reports_bp.route("/reports/export/csv")
@login_required
def export_csv():

    output = io.StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Date",
        "Revenue",
        "Visitors",
        "Orders",
        "Profit"
    ])

    metrics = Metric.query.filter_by(
        tenant_id=current_user.id
    ).order_by(
        Metric.timestamp.desc()
    ).all()

    for metric in metrics:

        writer.writerow([

            metric.timestamp.strftime("%d-%m-%Y"),

            metric.revenue,

            metric.visitors,

            metric.orders,

            metric.profit

        ])

    mem = io.BytesIO()

    mem.write(output.getvalue().encode("utf-8"))

    mem.seek(0)

    output.close()

    return send_file(

        mem,

        as_attachment=True,

        download_name="BuildHub_Report.csv",

        mimetype="text/csv"

    )